#!/usr/bin/env python3

# Copyright 2020 Jachym Cepicky (jachym.cepicky opengeolabs.cz)
#
# Permission is hereby granted, free of charge, to any person obtaining a copy of
# this software and associated documentation files (the "Software"), to deal in
# the Software without restriction, including without limitation the rights to
# use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies
# of the Software, and to permit persons to whom the Software is furnished to do
# so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

"""
Purpose:

    Convert input XLSX file into output CSV with two new attributes:
    lat/lon is also added as x, y

    X, Y is by default using Krovak projection  S-JTSK with transformation
    parameters used for Czech republic, as defined at https://epsg.io/5514-5239

Usage:

    see --help

Dependencies:

    Install with `pip3 install -r requirements.txt`

    * pyproj
    * openpyxl

Author:
    Jachym Cepicky <jachym.cepicky opengeolabs cz>
    http://opengeolabs.cz
"""

from openpyxl import load_workbook
import pyproj
import csv


def print_sheets(xlsx):
    """Just print content of the file, sheets and attributes
    :param xlsx: name of the input file
    """

    wb = load_workbook(xlsx)
    names = wb.get_sheet_names()

    for sheet_name in names:
        sheet = wb.get_sheet_by_name(sheet_name)
        vals = list(sheet.values)
        print(sheet_name)
        for val in vals[0]:
            print("\t{}".format(val))


def xlsx2csv(xlsx, output, sheetname, lon, lat):
    """Convert input file to output file

    :param xlsx: input file name
    :param output: output file name
    :param sheetname: name of the sheet in the table
    :param lon: name of the Longitude attribute
    :param lat: name of the Langitude attribute
    """

    wb = load_workbook(xlsx)
    sheet = wb.get_sheet_by_name(sheetname)

    vals = list(sheet.values)
    head = [v for v in vals[0]] + ["x", "y"]
    lon_idx = head.index(lon)
    lat_idx = head.index(lat)

    krovak = pyproj.Proj("""+proj=krovak +lat_0=49.5 +lon_0=24.83333333333333
            +alpha=30.28813972222222 +k=0.9999 +x_0=0 +y_0=0 +ellps=bessel
            +towgs84=572.213,85.334,461.94,-4.9732,-1.529,-5.2484,3.5378
            +units=m +no_defs""")
    wgs84 = pyproj.Proj("+init=epsg:4326")

    with open(output, "w") as output_file:

        writer = csv.writer(output_file)
        writer.writerow(head)

        for record in vals[1:]:
            lon = float(record[lon_idx])
            lat = float(record[lat_idx])
            x, y = pyproj.transform(wgs84, krovak, lon, lat)

            new_record = [r for r in record] + [x, y]
            writer.writerow(new_record)


def _get_args():
    import argparse

    parser = argparse.ArgumentParser(description='''XLSX to CSV with WGS84
            coordinates -> S-JTSK''')
    parser.add_argument('input', metavar='XLSZ', type=str,
                        help='input XLSX file')
    parser.add_argument('output', metavar="CSV", type=str,
                        help='output CSV')
    parser.add_argument('-l', action="store_true",
                        help='List sheets and names')
    parser.add_argument('--sheet', metavar="NAME", type=str,
                        help='Sheet name to convert')
    parser.add_argument('--loncolumn', metavar="NAME", type=str,
                        help='Name of the LONGITUTED column')
    parser.add_argument('--latcolumn', metavar="NAME", type=str,
                        help='Name of the LANGITUDE column')

    args = parser.parse_args()
    return args


def main():

    args = _get_args()

    if args.l:
        print_sheets(args.input)
    else:
        xlsx2csv(args.input, args.output, args.sheet,
                 args.loncolumn, args.latcolumn)


if __name__ == "__main__":
    main()
